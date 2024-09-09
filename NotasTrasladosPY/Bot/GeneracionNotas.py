from os.path import join, exists
from os import mkdir
import pandas as pd
import datetime as dt
import openpyxl as oxl
import os
import time

# Obtén la ruta del directorio actual donde se encuentra el archivo
current_dir = os.path.dirname(os.path.abspath(__file__))

# Mueve un nivel hacia arriba
workfolder_path = os.path.dirname(current_dir)

templates_path = join(workfolder_path, "templates", 'Plantilla_PDF_Nota.xlsx')

ProductosDB = os.path.join(workfolder_path, 'database', 'Productos Conciliaciones.xlsx')

df = pd.read_excel(ProductosDB, sheet_name='Hoja1')

# Inicializar el diccionario
products = {}

# Procesar cada columna y crear el diccionario
for column in df.columns:
    # Eliminar filas con valores NaN y convertir a una tupla
    products[column] = tuple(df[column].dropna().values)

# Unir los productos de todas las categorías en una sola cadena
all_products = []
for key in ['OTROS']:
    all_products.extend(products.get(key, []))

# Unir todos los productos en una sola cadena separada por '|'
product_str = '|'.join(all_products)

# Nombre de la carpeta donde serán almacenados los reportes generados.
folder_name = 'notas'

global today

# Ruta al reporte de notas (carpeta de OneDrive)
report_path = os.path.join(workfolder_path, 'Onedrive', 'reporte de notas.xls')

# Declaración de una variable para la fecha de hoy
today = dt.datetime.today().strftime("%d-%m-%Y")

# Crear la carpeta por fecha.
folder = join(workfolder_path, "reports", today)
# Si la carpeta por fecha no existe la crea.
if not exists(folder):
    mkdir(folder)

# Crear carpeta para notas
folder = join(folder, folder_name)
if not exists(folder):
    mkdir(folder)

# Lectura del reporte de la hoja de reporte de notas.
with open(report_path, mode='rb') as fp:
    df = pd.read_excel(fp, sheet_name='Hoja 1', dtype={'Nro Caso': str})

# Filtrar por columna 'Nro Caso' diferente a 'canal' o 'CANAL'
df_filtered = df[~df['Nro Caso'].str.contains('CANAL|canal', na=False)]

# Crear un Dataframe para productos diferentes a REC EPM EN LINEA o PAGO CONFAMA
df_filtered = df_filtered[df_filtered['Producto'].str.contains(product_str, na=False)]
df_filtered.reset_index(drop=True, inplace=True)

# Plantilla de Nota
# Workbook de la plantilla nota
workbook = oxl.load_workbook(templates_path)
# Seleccionar la primera hoja como worksheet
worksheet = workbook.active

cells = {
    "no": ("Id. Nota", 7),
    "tipo": ("Naturaleza", 8),
    "cc": ("Responsable", 9),
    "oficina": ("Oficina", 10),
    "centro": ("Oficina", 11),
    "valor": ("Valor", 12),
    "producto": ("Producto", 13),
    "caso": ("Nro Caso", 14),
    "obs": ("Observaciones", 16)
}

# Iterar sobre df_filtered para extraer los valores y escribirlos en un archivo basado en
# la plantilla de template
for index, row in df_filtered.iterrows():
    # Insertar la fecha de la nota "fecha": (today, 5)
    worksheet.cell(5, 3, f'{today}')
    for col_name, _row in cells.values():
        if col_name == 'Id. Nota':
            filename = row[col_name]
            worksheet.cell(row=_row, column=3, value=row[col_name])
        elif col_name == 'Valor':
            worksheet.cell(row=_row, column=3, value=f'${row[col_name]:,}')
        elif col_name == 'Oficina':
            if _row == 10:
                worksheet.cell(row=_row, column=3, value=row[col_name].split('|')[1])
            if _row == 11:
                worksheet.cell(row=_row, column=3, value=row[col_name].split('|')[0])
        else:
            worksheet.cell(row=_row, column=3, value=row[col_name])
    workbook = oxl.load_workbook(templates_path) #Para volver abrir el workbook y cogerla ahi mismo
    workbook.save(join(folder, f'{filename}_{today}.xlsx'))
    print(f'{filename}_{today} saved.')
workbook.close()
time.sleep(15)
print("All notas generated")