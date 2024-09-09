from os.path import join, exists
from os import mkdir
import os
import pandas as pd
import datetime as dt
import openpyxl as oxl
from openpyxl.styles import Border, Side

# Obtén la ruta del directorio actual donde se encuentra el archivo
current_dir = os.path.dirname(os.path.abspath(__file__))

# Mueve un nivel hacia arriba
workfolder_path = os.path.dirname(current_dir)

ProductosDB = os.path.join(workfolder_path, 'database', 'Productos Conciliaciones.xlsx')

df = pd.read_excel(ProductosDB, sheet_name='Hoja1')

# Inicializar el diccionario
products = {}

# Procesar cada columna y crear el diccionario
for column in df.columns:
    # Eliminar filas con valores NaN y convertir a una tupla
    products[column] = tuple(df[column].dropna().values)

# Unir los productos de todas las categorías en una sola cadena
all_products = []
for key in ['COMFAMA']:
    all_products.extend(products.get(key, []))

# Unir todos los productos en una sola cadena separada por '|'
comfama_products = '|'.join(all_products)

# Nombre de los reportes xls generados.
filename = 'Consolidado_COMFAMA'

# Nombre de la carpeta donde serán almacenados los reportes generados.
folder_name = 'comfama'

global today

# Ruta al reporte de notas (carpeta de OneDrive)
report_path = os.path.join(workfolder_path, 'Onedrive', 'reporte de notas.xls')

# Ruta al archivo plantilla PDF_Comfama
template_path = join(workfolder_path, "templates", 'Plantilla_PDF_Comfama.xlsx')

# Declaración de una variable para la fecha de hoy
today = dt.datetime.today().strftime("%d-%m-%Y")

# Crear la carpeta por fecha si no existe
folder = join(workfolder_path, "reports", today)
if not exists(folder):
    mkdir(folder)

# Crear carpeta para consolidado
folder = join(folder, folder_name)
if not exists(folder):
    mkdir(folder)

# Lectura del reporte de la hoja de reporte de notas.
with open(report_path, mode='rb') as fp:
    df = pd.read_excel(fp, sheet_name='Hoja 1', dtype={'Nro Caso': str})

# Filtrar por columna 'Nro Caso' diferente a 'canal' o 'CANAL'
df_filtered = df[~df['Nro Caso'].str.contains('CANAL|canal', na=False)]

df_filtered = df_filtered[df_filtered['Producto'].str.contains(comfama_products, na=False)]
df_filtered.reset_index(drop=True, inplace=True)

# Plantilla de consolidado COMFAMA
# Workbook de la plantilla consolidado COMFAMA
workbook = oxl.load_workbook(template_path)
# Seleccionar la primera hoja como worksheet
worksheet = workbook.active

cells = {
    "id": ("Id. Nota", 2),
    "centro": ("Oficina", 3),
    "naturaleza": ("Naturaleza", 4),
    "caso": ("Nro Caso", 5),
    "producto": ("Producto", 6),
    "responsable": ("Responsable", 7),
    "obs": ("Observaciones", 8),
    "valor": ("Valor", 9),
    "aliado": ("COMFAMA", 10)
}

# Estilo de las celdas
border_style = Border(
    left=Side(border_style='thin', color='FF000000'),
    right=Side(border_style='thin', color='FF000000'),
    top=Side(border_style='thin', color='FF000000'),
    bottom=Side(border_style='thin', color='FF000000'),
)

# Rango de la plantilla
min_row = 7
min_col = 2
max_col = 10
max_row = min_row

if df_filtered.shape[0] > 1:
    max_row = min_row + df_filtered.shape[0] - 1
    for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = border_style

total = 0
worksheet.cell(1, 6, f'{today}')

for index, row in df_filtered.iterrows():
    for col_name, col in cells.values():
        if col_name == 'COMFAMA':
            worksheet.cell(row=index + min_row, column=col, value=col_name)
        elif col_name == 'Valor':
            total += int(row[col_name])
            worksheet.cell(row=index + min_row, column=col, value=f'${row[col_name]:,}')
        else:
            worksheet.cell(row=index + min_row, column=col, value=row[col_name])

total_cell = worksheet.cell(row=max_row + 1, column=cells.get('valor')[1], value=f'${total:,}')
total_cell.border = border_style

workbook.save(join(folder, f'{filename}_{today}.xlsx'))
workbook.close()

print(f'{filename}_{today} saved.')