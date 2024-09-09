import pandas as pd
import openpyxl as oxl
import os
import datetime as dt
import os


# Obtén la ruta del directorio actual donde se encuentra el archivo
current_dir = os.path.dirname(os.path.abspath(__file__))

# Mueve un nivel hacia arriba
workfolder_path = os.path.dirname(current_dir)

def load_database(workfolder_path, db_name):
    """Carga la base de datos desde el archivo Excel."""
    db_path = os.path.join(workfolder_path, 'database', db_name)
    try:
        with open(db_path, mode='rb') as fp:
            df = pd.read_excel(fp, sheet_name='PRODUCTOS', engine='openpyxl')
        print("Base de datos cargada correctamente.")
        return df
    except Exception as e:
        print(f"Error al cargar la base de datos: {e}")
        return None

def main():
    db_name = os.path.join(workfolder_path,'database', 'DB Producto Conciliaciones V2.xlsx')
    ProductosDB = os.path.join(workfolder_path, 'database', 'Productos Conciliaciones.xlsx')
    
    control_template = 'Plantilla_ControlOp.xlsx'
    template_path = os.path.join(workfolder_path, "templates", control_template)
    
    df = pd.read_excel(ProductosDB, sheet_name='Hoja1')

    # Inicializar el diccionario
    products = {}

    # Procesar cada columna y crear el diccionario
    for column in df.columns:
        # Eliminar filas con valores NaN y convertir a una tupla
        products[column] = tuple(df[column].dropna().values)

    # Unir los productos de todas las categorías en una sola cadena
    all_products = []
    for key in ['EPM', 'COMFAMA', 'OTROS']:
        all_products.extend(products.get(key, []))

    # Unir todos los productos en una sola cadena separada por '|'
    products = '|'.join(all_products)

    # Nombre del reporte xls generados.
    filename = 'CONSOLIDADO'

    # Ruta al reporte de notas (carpeta de OneDrive)
    report_path = os.path.join(workfolder_path, 'Onedrive', 'reporte de notas.xls')

    # Declaración de una variable para la fecha de hoy
    today = dt.datetime.today().strftime("%d-%m-%Y")
    # Generar el nombre del reporte con la fecha de hoy
    filename = f'{filename} {today}'

    # Crear la carpeta por fecha.
    folder = os.path.join(workfolder_path, "reports", today)
    # Si la carpeta por fecha no existe la crea.
    if not os.path.exists(folder):
        os.mkdir(folder)
        print(f"Creada la carpeta de reportes en: {folder}")
    else:
        print(f"La carpeta de reportes ya existe en: {folder}")

    # Lectura del reporte de la hoja de reporte de notas.
    try:
        df = pd.read_excel(report_path, sheet_name='Hoja 1', dtype={'Nro Caso': str})
        print("Reporte de notas leído correctamente.")
    except Exception as e:
        print(f"Error al leer el reporte de notas: {e}")

    # Filtrar por columna 'Nro Caso' diferente a 'canal' o 'CANAL'
    df_filtered = df[~df['Nro Caso'].str.contains('CANAL|canal', na=False)]
    # Filtrar dataframe por productos que esten en la Base de datos
    df_filtered = df_filtered[df_filtered['Producto'].str.contains(products, na=False)]
    # Reiniciar el indice del dataframe
    df_filtered.reset_index(drop=True, inplace=True)

    try:
        df_db = load_database(workfolder_path, db_name)
    except Exception as e:
        print(f"Error al cargar la base de datos: {e}")
        return

    # Plantilla de Consolidado control
    try:
        workbook = oxl.load_workbook(template_path)
        worksheet = workbook.active

        cells = {
            "fecha": ("Fecha", 1),
            "auxiliar": ("AUXILIAR", 2),
            "id": ("Id. Nota", 3),
            "centro": ("Oficina", 4),
            "naturaleza": ("Naturaleza", 5),
            "producto": ("Producto", 6),
            "responsable": ("Responsable", 7),
            "valor": ("Valor", 8),
            "aliado": ("ALIADO", 9),
            "obs": ("Observaciones", 10),
            "caso": ("Nro Caso", 11)
        }

        min_row = 3
        min_col = 1
        max_col = 11
        max_row = min_row + df_filtered.shape[0] - 1 if df_filtered.shape[0] > 1 else min_row

        border_style = oxl.styles.Border(
            left=oxl.styles.Side(border_style='thin', color='FF000000'),
            right=oxl.styles.Side(border_style='thin', color='FF000000'),
            top=oxl.styles.Side(border_style='thin', color='FF000000'),
            bottom=oxl.styles.Side(border_style='thin', color='FF000000'),
        )

        for row in worksheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = border_style

        for index, row in df_filtered.iterrows():
            for col_name, col in cells.values():
                if col_name == 'Fecha':
                    worksheet.cell(row=index + min_row, column=col, value=today)
                elif col_name == 'AUXILIAR':
                    auxiliar = df_db[df_db.PRODUCTO == row['Producto']].AUXILIAR.to_list()
                    worksheet.cell(row=index + min_row, column=col, value=auxiliar[0] if auxiliar else '')
                elif col_name == 'ALIADO':
                    aliado = df_db[df_db.PRODUCTO == row['Producto']].ALIADO.to_list()
                    worksheet.cell(row=index + min_row, column=col, value=aliado[0] if aliado else '')
                else:
                    worksheet.cell(row=index + min_row, column=col, value=row[col_name])

        workbook.save(os.path.join(folder, f'{filename}.xlsx'))
        workbook.close()
        print(f'{filename} Consolidado general guardado correctamente.')
    except Exception as e:
        print(f"Error al generar el consolidado general: {e}")

if __name__ == "__main__":
    main()
